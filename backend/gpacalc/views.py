import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from scheduler.models import Course, CourseEvent
from .models import CourseGrade, AssessmentGrade, GpaCalcProgress, GradingScheme, AssessmentWeightage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

logger = logging.getLogger(__name__)
app_name = "GPA-Calculator"
@require_http_methods(["GET", "POST"])
@csrf_exempt
def get_offered_terms(request):
    """
    API: Get all available terms
    
    Returns:
        JSON: Array of term strings ["Fall 2025", "Winter 2026", ...]
    
    Frontend usage:
    - Call to populate term dropdown in UI
    """
    data = json.loads(request.body) if request.body else {}
    # Added has_events flag because for scheduler app, we don't need to have courses with events
    # And we won't have events when students will use scheduler app because by then we wont have 
    # Course outlines to extract events from.
    has_events = data.get("has_events", True) 
    print(f"Fetching offered terms with has_events={has_events}")
    terms = (
            Course.objects
            .filter(has_events=has_events)
            .values_list("offered_term", flat=True)
            .distinct()
            .order_by("offered_term")
        )
    terms_list = list(terms)  # Execute the query
    response = JsonResponse(terms_list, safe=False)
    
    return response

@require_POST
@csrf_exempt
def get_course_types(request):
    """
    API: Get course types for a given term
    
    Request:
        JSON: {"offered_term": "Fall 2025"}
    
    Returns:
        JSON: Array of course types ["CS", "MATH", ...]
    
    Frontend usage:
    - Call when user selects a term to populate the course type dropdown
    """
    
    data = json.loads(request.body)
    cterm = data.get("offered_term")
    # Added has_events flag because for scheduler app, we don't need to have courses with events
    # And we won't have events when students will use scheduler app because by then we wont have 
    # Course outlines to extract events from.
    has_events = data.get("has_events", True) 
    try:
        types = (
            Course.objects
            .filter(has_events=has_events, offered_term=cterm)
            .values_list("course_type", flat=True)
            .distinct()
            .order_by("course_type")
        )
        types_list = list(types)
    except Exception as e:
        logger.error(f"Fetching course types failed: {e}")
        types_list = []
    response = JsonResponse(types_list, safe=False)
    return response

@require_POST
@csrf_exempt
def get_course_codes(request):
    """
    API: Get course codes for a given course type and term
    
    Request:
        JSON: {
            "course_type": "CS",
            "offered_term": "Fall 2025"
        }
    
    Returns:
        JSON: Array of course codes ["101", "202", ...]
    
    Frontend usage:
    - Call when user selects a course type to populate the course code dropdown
    """
    data = json.loads(request.body)
    # Added has_events flag because for scheduler app, we don't need to have courses with events
    # And we won't have events when students will use scheduler app because by then we wont have 
    # Course outlines to extract events from.
    has_events = data.get("has_events", True) 
    cterm = data.get("offered_term")
    ctype = data.get("course_type")
    try:
        codes = (
            Course.objects
            .filter(course_type=ctype, offered_term=cterm, has_events=has_events) #2025-08-01 OPTIMIZED: Uses has_events boolean field instead of JOIN Performance: ~5-10ms instead of 120ms
            .values_list("course_code", flat=True)
            .distinct()
            .order_by("course_code")
        )
    except Exception as e:
        logger.error(f"Fetching course codes failed: {e}")
        codes = []
    return JsonResponse(list(codes), safe=False)

@require_POST
@csrf_exempt
def get_section_numbers(request):
    """
    API: Get section numbers for a given course type, code, and term
    
    Request:
        JSON: {
            "course_type": "CS",
            "course_code": "101",
            "offered_term": "Fall 2025"
        }
    
    Returns:
        JSON: Array of section numbers ["001", "002", ...]
    
    Frontend usage:
    - Call when user selects a course code to populate the section dropdown
    """
    data = json.loads(request.body)
     # Added has_events flag because for scheduler app, we don't need to have courses with events
    # And we won't have events when students will use scheduler app because by then we wont have 
    # Course outlines to extract events from.
    has_events = data.get("has_events", True) 
    ctype = data.get("course_type")
    ccode = data.get("course_code")
    cterm = data.get("offered_term")
    secs = (
        Course.objects
        .filter(course_type=ctype, offered_term=cterm, course_code=ccode, has_events=has_events) #2025-08-01 OPTIMIZED: Uses has_events boolean field instead of JOIN Performance: ~5-10ms instead of 120ms
        .values_list("section_number", flat=True)
        .distinct()
        .order_by("section_number")
    )
    return JsonResponse(list(secs), safe=False)

@require_POST
@csrf_exempt
def get_course_events(request):
    """
    API: Get assessment events for a specific course section
    
    Request:
        JSON: {
            "course_type": "CS",
            "course_code": "101",
            "section_number": "001",
            "offered_term": "Fall 2025"
        }
    
    Returns:
        JSON: Array of event objects
        [
            {
                "id": 12,
                "event_type": "Midterm",
                "weightage": "30"
            },
            ...
        ]
    
    Frontend usage:
    - Call when user selects a section to display assessment components
    - Display each event with its weight and an input field for the achieved grade
    - Store the event_id with each input field to submit with calculations
    """
    data = json.loads(request.body)
    ctype = data.get("course_type")
    code = data.get("course_code")
    csn = data.get("section_number")
    cterm = data.get("offered_term")
    user = request.user if request.user.is_authenticated else None
    evs = CourseEvent.objects.filter(
        course__course_type=ctype,
        course__course_code=code,
        course__section_number=csn,
        course__offered_term=cterm
    ).values("id", "event_type", "weightage","event_date").order_by("event_date")
    return JsonResponse(list(evs), safe=False)

@require_POST
@csrf_exempt
def calculate_gpa(request):
    """
    API: Calculate GPA based on course selections and grades, considering multiple grading schemes
    
    Request:
        JSON: {
            "offered_term": "Fall 2025",
            "courses": [
                {
                    "course_type": "CS",
                    "course_code": "101",
                    "section_number": "001",
                    "assessments": [
                        {"event_id": 12, "achieved": 87.5},
                        ...
                    ]
                },
                ...
            ]
        }
    
    Returns:
        JSON: {
            "combinations": [
                {
                    "scheme_id": "combo_0",
                    "scheme_name": "CS 101: Default + MATH 135: Scheme A",
                    "per_course": [
                        {
                            "course": "CS101-001",
                            "final_percentage": 85.0,
                            "letter_grade": "A",
                            "gpa_value": 4.0,
                            "credits": 0.5,
                            "scheme_name": "Default"
                        },
                        ...
                    ],
                    "overall_gpa": 3.7,
                    "overall_final_percentage": 82.5,
                    "schemes_used": ["CS 101: Default", "MATH 135: Scheme A"]
                },
                ...
            ],
            "best_combination": {...},  // The combination with the highest GPA
            "individual_schemes": [
                {
                    "course": "CS101-001",
                    "course_type": "CS",
                    "course_code": "101",
                    "section_number": "001",
                    "scheme_id": "default",
                    "scheme_name": "Default",
                    "scheme_description": "Default grading scheme",
                    "weightages": {"Midterm": "30%", "Final": "50%"},
                    "final_percentage": 85.0,
                    "letter_grade": "A",
                    "gpa_value": 4.0
                },
                ...
            ],
            "total_credit": 2.5,
            "courses": [...],  // Cleaned course data
            "offered_term": "Fall 2025"
        }
    
    Frontend usage:
    - Call after user enters all grades to calculate GPA
    - Display the overall GPA for the best combination of grading schemes
    - Allow users to explore different grading scheme combinations
    - Show how each grading scheme affects individual course grades
    - For logged-in users, this data is automatically saved
    """
    
    data = json.loads(request.body)
    offered_term = data.get("offered_term")
    
    # Store course objects and their schemes for permutation calculation
    course_schemes = []
    total_credit = 0
    
    # First pass: Gather all courses and their available schemes
    for c in data.get("courses", []):
        course_obj = Course.objects.get(
            course_type=c["course_type"],
            course_code=c["course_code"],
            section_number=c["section_number"],
            offered_term=offered_term,
        )
        
        total_credit += float(course_obj.credits)
        
        # Get all grading schemes for this course
        grading_schemes = list(GradingScheme.objects.filter(course=course_obj))
        
        # If no schemes exist, create a virtual default one
        if not grading_schemes:
            course_schemes.append({
                "course": course_obj,
                "assessments": c.get("assessments", []),
                "schemes": ["default"]  # Use special marker for default scheme
            })
        else:
            course_schemes.append({
                "course": course_obj,
                "assessments": c.get("assessments", []),
                "schemes": grading_schemes
            })
    
    # Generate all possible scheme combinations
    def get_scheme_combinations(courses_with_schemes, current=None, index=0):
        if current is None:
            current = []
        
        if index >= len(courses_with_schemes):
            return [current]
        
        result = []
        course_data = courses_with_schemes[index]
        for scheme in course_data["schemes"]:
            new_current = current + [(course_data["course"], course_data["assessments"], scheme)]
            result.extend(get_scheme_combinations(courses_with_schemes, new_current, index+1))
        
        return result
    
    all_combinations = get_scheme_combinations(course_schemes)
    
    # Calculate GPA for each combination
    combination_results = []
    
    for combo in all_combinations:
        combo_total_points = 0
        combo_total_weighted_percentage = 0
        combo_courses = []
        scheme_names = []
        
        for course_obj, assessments, scheme in combo:
            if scheme == "default":
                # Use default weightages
                course_grade = calculate_for_default_scheme(course_obj, assessments)
                scheme_name = "Default"
            else:
                # Use scheme weightages
                course_grade = calculate_for_scheme(course_obj, scheme, assessments)
                scheme_name = scheme.name
            
            scheme_names.append(f"{course_obj.course_type} {course_obj.course_code}: {scheme_name}")
            
            course_result = {
                "course": str(course_obj),
                "final_percentage": float(course_grade.final_percentage) if course_grade.final_percentage else 0,
                "letter_grade": course_grade.letter_grade or "N/A",
                "gpa_value": float(course_grade.gpa_value) if course_grade.gpa_value else 0,
                "credits": float(course_obj.credits),
                "scheme_name": scheme_name
            }
            
            combo_courses.append(course_result)
            
            if course_grade.gpa_value:
                combo_total_points += float(course_grade.gpa_value) * float(course_obj.credits)
                
            if course_grade.final_percentage:
                combo_total_weighted_percentage += float(course_grade.final_percentage) * float(course_obj.credits)
        
        # Calculate overall GPA for this combination
        combo_overall_gpa = round(combo_total_points / total_credit, 2) if total_credit > 0 else 0
        combo_overall_percentage = round(combo_total_weighted_percentage / total_credit, 2) if total_credit > 0 else 0
        
        combination_results.append({
            "scheme_id": f"combo_{len(combination_results)}",
            "scheme_name": " + ".join(scheme_names),
            "per_course": combo_courses,
            "overall_gpa": combo_overall_gpa,
            "overall_final_percentage": combo_overall_percentage,
            "schemes_used": scheme_names
        })
    
    # Find the best combination
    best_combo = max(
        combination_results, 
        key=lambda x: (x["overall_gpa"], x["overall_final_percentage"])
    ) if combination_results else {}
    
    # Also prepare individual scheme results for each course (for UI display)
    individual_scheme_results = []
    
    for course_data in course_schemes:
        course_obj = course_data["course"]
        assessments = course_data["assessments"]
        
        for scheme in course_data["schemes"]:
            if scheme == "default":
                course_grade = calculate_for_default_scheme(course_obj, assessments)
                scheme_name = "Default"
                scheme_id = "default"
                scheme_description = "Default grading scheme"
                weightages = {}
                
                # Get weightages for display
                for event in CourseEvent.objects.filter(course=course_obj):
                    if event.weightage:
                        weightages[event.event_type] = event.weightage
                
            else:
                course_grade = calculate_for_scheme(course_obj, scheme, assessments)
                scheme_name = scheme.name
                scheme_id = str(scheme.id)
                scheme_description = scheme.description or ""
                
                # Get weightages for display
                weightages = {}
                for w in AssessmentWeightage.objects.filter(grading_scheme=scheme):
                    weightages[w.course_event.event_type] = f"{w.weightage}%"
            
            individual_scheme_results.append({
                "course": str(course_obj),
                "course_type": course_obj.course_type,
                "course_code": course_obj.course_code,
                "section_number": course_obj.section_number,
                "scheme_id": scheme_id,
                "scheme_name": scheme_name,
                "scheme_description": scheme_description,
                "weightages": weightages,
                "final_percentage": float(course_grade.final_percentage) if course_grade.final_percentage else 0,
                "letter_grade": course_grade.letter_grade or "N/A",
                "gpa_value": float(course_grade.gpa_value) if course_grade.gpa_value else 0,
            })
    
    # Save the original assessments for progress tracking
    clean_courses = []
    for course in data.get("courses", []):
        clean_assessments = []
        for a in course.get("assessments", []):
            clean_assessments.append({
                "event_id": int(a.get("event_id")) if a.get("event_id") is not None else None,
                "achieved": float(a.get("achieved")) if a.get("achieved") is not None else None,
            })
        clean_courses.append({
            "course_type": course.get("course_type", ""),
            "course_code": course.get("course_code", ""),
            "section_number": course.get("section_number", ""),
            "assessments": clean_assessments,
        })

    result_data = {
        "combinations": combination_results,
        "best_combination": best_combo,
        "individual_schemes": individual_scheme_results,
        "total_credit": float(total_credit),
        "courses": clean_courses,
        "offered_term": data.get("offered_term", ""),
    }
    
    # Save progress for logged-in users
    if request.user.is_authenticated:
        GpaCalcProgress.objects.update_or_create(
            user=request.user,
            defaults={'data': result_data}
        )
    else:
        request.session['gpacalc_progress'] = result_data
    
    return JsonResponse(result_data)


def calculate_for_default_scheme(course, assessments):
    """
    Calculate grades using the default weightages from CourseEvent
    
    Args:
        course: Course object
        assessments: List of assessment data with event_id and achieved percentage
        
    Returns:
        CourseGrade object with calculated final percentage, letter grade and GPA value
    """
    cg = CourseGrade.objects.create(course=course)
    
    for a in assessments:
        try:
            ev = CourseEvent.objects.get(id=a["event_id"])
            # Parse weightage to decimal (strip % if present)
            raw_weight = ev.weightage or 0
            if isinstance(raw_weight, str) and raw_weight.endswith("%"):
                raw_weight = raw_weight.rstrip("%")
            try:
                weight = float(raw_weight)
            except Exception:
                weight = 0
                
            AssessmentGrade.objects.create(
                course_grade=cg,
                course_event=ev,
                weightage=weight,
                achieved_percentage=a["achieved"]
            )
        except Exception as e:
            continue  # Skip invalid assessments
    
    cg.calculate_from_assessments()
    return cg


def calculate_for_scheme(course, scheme, assessments):
    """
    Calculate grades using a specific grading scheme
    
    Args:
        course: Course object
        scheme: GradingScheme object
        assessments: List of assessment data with event_id and achieved percentage
        
    Returns:
        CourseGrade object with calculated final percentage, letter grade and GPA value
    """
    cg = CourseGrade.objects.create(course=course)
    
    for a in assessments:
        try:
            ev = CourseEvent.objects.get(id=a["event_id"])
            
            # Get weightage from the grading scheme
            try:
                aw = AssessmentWeightage.objects.get(grading_scheme=scheme, course_event=ev)
                weight = float(aw.weightage)
            except AssessmentWeightage.DoesNotExist:
                # Fall back to event's weightage
                raw_weight = ev.weightage or 0
                if isinstance(raw_weight, str) and raw_weight.endswith("%"):
                    raw_weight = raw_weight.rstrip("%")
                try:
                    weight = float(raw_weight)
                except Exception:
                    weight = 0
            
            AssessmentGrade.objects.create(
                course_grade=cg,
                course_event=ev,
                weightage=weight,
                achieved_percentage=a["achieved"]
            )
        except Exception as e:
            None
    
    cg.calculate_from_assessments()
    return cg


@require_GET
@csrf_exempt
def progress_export_excel(request):
    """
    API: Export GPA calculation results to Excel file, including multiple grading schemes
    
    Request: GET request (no parameters needed)
    
    Returns:
        Excel file as attachment with the following sheets:
        - Best Combination: Summary of the optimal grading scheme combination
        - All Combinations: Overview of all possible grading scheme combinations
        - Individual Schemes: How each scheme affects each course
        - Scheme Weightages: Detailed breakdown of assessment weights per scheme
        - Per-course sheets: One sheet per course with assessment details
    
    Frontend usage:
    - Provide an "Export to Excel" button that links to this endpoint
    - Uses the saved progress data (for logged-in users) or session data
    - Comprehensive report showing optimal grading scheme selection
    """
    
    # Get progress data
    if request.user.is_authenticated:
        try:
            progress = GpaCalcProgress.objects.get(user=request.user)
            data = progress.data
        except GpaCalcProgress.DoesNotExist:
            data = {}
    else:
        data = request.session.get('gpacalc_progress', {})
    
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Add Best Combination Sheet
    best_combo = data.get('best_combination', {})
    if best_combo:
        ws_best = wb.create_sheet(title="Best Combination")
        ws_best.append(["Overall GPA", best_combo.get('overall_gpa', 'N/A')])
        ws_best.append(["Overall Final %", best_combo.get('overall_final_percentage', 'N/A')])
        ws_best.append(["Total Credits", data.get('total_credit', 'N/A')])
        ws_best.append([])  # Blank row
        
        # Add schemes used
        ws_best.append(["Schemes Used:"])
        for scheme in best_combo.get('schemes_used', []):
            ws_best.append(["• " + scheme])
        
        ws_best.append([])  # Blank row
        ws_best.append(["Course", "Final %", "Letter Grade", "GPA Value", "Credits", "Scheme Used"])
        
        for c in best_combo.get('per_course', []):
            ws_best.append([
                c.get("course", ""),
                c.get("final_percentage", 'N/A'),
                c.get("letter_grade", ""),
                c.get("gpa_value", 'N/A'),
                c.get("credits", 'N/A'),
                c.get("scheme_name", "")
            ])
        
        # Format the sheet
        for col in range(1, 7):
            ws_best.column_dimensions[get_column_letter(col)].width = 18

    # 2. Add All Combinations Sheet
    combinations = data.get('combinations', [])
    if combinations:
        ws_all = wb.create_sheet(title="All Combinations")
        ws_all.append(["Combination", "Overall GPA", "Overall Final %", "Schemes Used"])
        
        for i, combo in enumerate(combinations):
            is_best = combo.get('scheme_id') == best_combo.get('scheme_id')
            row_num = i + 2  # +2 because we have a header row and Excel is 1-indexed
            
            ws_all.append([
                f"Combination {i+1}" + (" (Best)" if is_best else ""),
                combo.get('overall_gpa', 'N/A'),
                combo.get('overall_final_percentage', 'N/A'),
                ", ".join(combo.get('schemes_used', []))
            ])
            
            # Highlight the best combination
            if is_best:
                for col in range(1, 5):
                    cell = ws_all.cell(row=row_num, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                    )
        
        # Format the sheet
        for col in range(1, 5):
            ws_all.column_dimensions[get_column_letter(col)].width = 30

    # 3. Add Individual Schemes Sheet
    individual_schemes = data.get('individual_schemes', [])
    if individual_schemes:
        ws_schemes = wb.create_sheet(title="Individual Schemes")
        ws_schemes.append(["Course", "Scheme", "GPA Value", "Final %", "Letter Grade"])
        
        for scheme in individual_schemes:
            ws_schemes.append([
                scheme.get('course', ''),
                scheme.get('scheme_name', ''),
                scheme.get('gpa_value', 'N/A'),
                scheme.get('final_percentage', 'N/A'),
                scheme.get('letter_grade', '')
            ])
        
        # Format the sheet
        for col in range(1, 6):
            ws_schemes.column_dimensions[get_column_letter(col)].width = 18

    # 4. Add Scheme Weightages Sheet
    ws_weights = wb.create_sheet(title="Scheme Weightages")
    ws_weights.append(["Course", "Scheme", "Assessment", "Weight"])
    
    row = 2
    for scheme in individual_schemes:
        course = scheme.get('course', '')
        scheme_name = scheme.get('scheme_name', '')
        weightages = scheme.get('weightages', {})
        
        if weightages:
            first_row = True
            for assessment, weight in weightages.items():
                if first_row:
                    ws_weights.append([course, scheme_name, assessment, weight])
                    first_row = False
                else:
                    ws_weights.append(['', '', assessment, weight])
                row += 1
        else:
            ws_weights.append([course, scheme_name, 'No weightages available', ''])
            row += 1
    
    # Format the sheet
    for col in range(1, 5):
        ws_weights.column_dimensions[get_column_letter(col)].width = 20

    # 5. Add individual course sheets with assessment data
    for course in data.get('courses', []):
        sheet_name = f"{course.get('course_type', '')}-{course.get('course_code', '')}-{course.get('section_number', '')}"
        sheet_name = sheet_name[:31]  # Excel sheet name length limit
        
        # Skip if a sheet with this name already exists
        if sheet_name in [ws.title for ws in wb.worksheets]:
            sheet_name = f"{sheet_name[:27]}-{len(wb.worksheets)}"
        
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Term', 'Assessment', 'Date', 'Weightage', 'Achieved', 'Achieved %'])
        section_name = f"{course.get('course_type', '')}*{course.get('course_code', '')}*{course.get('section_number', '')}"
        course_temp = Course.objects.filter(section_name=section_name)
        course_id = course_temp.first().course_id if course else None
        for assessment in course.get('assessments', []):
            event_id = assessment.get('event_id', '')
            try:
                event = CourseEvent.objects.get(id=event_id)
                description = event.event_type
                event_date = event.event_date
                per_course = data.get('best_combination', {}).get('per_course', {})
                for course in per_course:
                    if course.get('course') == section_name: 
                        scheme_name= course.get('scheme_name', '')
                grading_scheme = GradingScheme.objects.filter(
                    course=course_id,name = scheme_name).first()
                if grading_scheme:
                    scheme_id = grading_scheme.id
                weightage = AssessmentWeightage.objects.filter(
                    grading_scheme__id=scheme_id,
                    course_event__id=event_id
                ).first()
                weightage = weightage.weightage if weightage else event.weightage
            except Exception:
                description = ''
                event_date = ''
                weightage = ''
            
            achieved = assessment.get('achieved', '')
            
            # Convert weightage to float if possible
            try:
                weightage_val = float(str(weightage).strip('%')) if weightage not in (None, '') else None
            except Exception:
                weightage_val = None
            
            # Achieved as float
            try:
                achieved_val = float(achieved) if achieved not in (None, '') else None
            except Exception:
                achieved_val = None
            
            # Achieved % as float (percentage points contribution)
            try:
                achieved_pct = (achieved_val * weightage_val) / 100 if achieved_val is not None and weightage_val is not None else None
            except Exception:
                achieved_pct = None
            
            ws.append([
                data.get('offered_term', ''),
                description,
                event_date,
                weightage_val,
                achieved_val,
                achieved_pct
            ])
        
        # Set percentage format for Achieved % column
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '0.00'  # Just show as decimal
        
        # Format the sheet
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # Save and return the Excel file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="gpacalc_progress.xlsx"'
    
    return response


@require_GET
def get_user_progress(request):
    """
    API: Get the user's saved GPA calculation progress
    
    Returns:
        JSON: The user's most recent GPA calculation data or empty object if none exists
    
    Frontend usage:
    - Call when loading the GPA calculator page to restore previous state
    - Used to prefill form fields and display previously calculated results
    - Returns empty object if no previous calculations exist
    """
    if request.user.is_authenticated:
        try:
            progress = GpaCalcProgress.objects.get(user=request.user)
            return JsonResponse(progress.data)
        except GpaCalcProgress.DoesNotExist:
            return JsonResponse({})
    else:
        return JsonResponse(request.session.get('gpacalc_progress', {}))